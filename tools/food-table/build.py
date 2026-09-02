#!/usr/bin/env python3
"""
Builds the reference food table the backend imports.

The Swiss Food Composition Database is published as one workbook per language,
each with the same rows and the same ids. This joins the four of them into two
CSVs the backend reads at boot:

  backend/src/main/resources/food/foods.csv   — one row per food, macros
  backend/src/main/resources/food/names.csv   — every name and synonym, by locale

Run it whenever the source publishes a new version. Nothing here is edited by
hand: the CSVs are generated, the backend re-imports when their checksum
changes, and the diff of a refresh is reviewable.

    docker run --rm -v "$PWD:/repo" -w /repo python:3.12-slim \
        sh -lc 'pip install -q openpyxl requests && python tools/food-table/build.py'

Add --from-cache to reuse workbooks already downloaded into tools/food-table/cache.

Source: Swiss Food Composition Database, Federal Food Safety and Veterinary
Office (FSVO). Free to use, including commercially, with the source named — see
tools/food-table/SOURCE.md, which is the file to keep in step with their terms.
"""

import argparse
import csv
import pathlib
import re
import sys
import unicodedata

CACHE = pathlib.Path(__file__).parent / "cache"
OUT = pathlib.Path(__file__).parents[2] / "backend/src/main/resources/food"

# One workbook per language, same ids in every one.
WORKBOOKS = {
    "en": "https://naehrwertdaten.ch/wp-content/uploads/2026/07/Swiss_food_composition_database.xlsx",
    "fr": "https://naehrwertdaten.ch/wp-content/uploads/2026/07/Base_de_donnees_suisse_des_valeurs_nutritives.xlsx",
    "de": "https://naehrwertdaten.ch/wp-content/uploads/2026/07/Schweizer_Nahrwertdatenbank.xlsx",
}

# The English column headers we read, mapped to the names the CSV uses. The
# workbook repeats "Derivation of value" and "Source" after every value column;
# both are dropped, and the header row is what tells us which is which.
MACROS = {
    "Energy, kilocalories (kcal)": "kcal",
    "Protein (g)": "protein_g",
    "Carbohydrates, available (g)": "carbs_g",
    "Fat, total (g)": "fat_g",
    "Dietary fibres (g)": "fibre_g",
    "Sugars (g)": "sugars_g",
    "Salt (NaCl) (g)": "salt_g",
}

# What the detail panel shows. Short codes rather than the workbook's prose, so
# a column rename upstream is a change here and nowhere else.
MICRONUTRIENTS = {
    "Sodium (Na) (mg)": ("sodium", "mg"),
    "Potassium (K) (mg)": ("potassium", "mg"),
    "Calcium (Ca) (mg)": ("calcium", "mg"),
    "Magnesium (Mg) (mg)": ("magnesium", "mg"),
    "Phosphorus (P) (mg)": ("phosphorus", "mg"),
    "Iron (Fe) (mg)": ("iron", "mg"),
    "Zinc (Zn) (mg)": ("zinc", "mg"),
    "Selenium (Se) (µg)": ("selenium", "ug"),
    "Iodide (I) (µg)": ("iodine", "ug"),
    "Vitamin A activity, RAE (µg)": ("vitamin_a", "ug"),
    "Vitamin B1 (thiamine) (mg)": ("vitamin_b1", "mg"),
    "Vitamin B2 (riboflavin) (mg)": ("vitamin_b2", "mg"),
    "Vitamin B6 (pyridoxine) (mg)": ("vitamin_b6", "mg"),
    "Vitamin B12 (cobalamin) (µg)": ("vitamin_b12", "ug"),
    "Niacin (mg)": ("niacin", "mg"),
    "Folate (µg)": ("folate", "ug"),
    "Pantothenic acid (mg)": ("pantothenic_acid", "mg"),
    "Vitamin C (ascorbic acid) (mg)": ("vitamin_c", "mg"),
    "Vitamin D (calciferol) (µg)": ("vitamin_d", "ug"),
    "Vitamin E (α-tocopherol) (mg)": ("vitamin_e", "mg"),
    "Cholesterol (mg)": ("cholesterol", "mg"),
    "Fatty acids, saturated (g)": ("saturated_fat", "g"),
}

# The workbook's category, mapped to the aisle the shopping list groups by.
# The full path is consulted before the top level, because a shop does not
# arrange itself the way a composition table does: butter and oil share a
# category here and two different aisles there.
AISLES = {
    # Full paths first: a shop does not arrange itself the way a composition
    # table does. Butter and oil share a category here and two aisles there,
    # and a potato is produce however much starch it has.
    "Fats and oils/Fats": "DAIRY",
    "Fats and oils/Cream": "DAIRY",
    "Cereal products, pulses and potatoes/Potatoes and other starchy tubers": "PRODUCE",
    "Cereal products, pulses and potatoes/Dough": "BAKERY",
    "Bread, flakes and breakfast cereals/Flakes, bran and germs": "GROCERY",
    "Bread, flakes and breakfast cereals/Muesli mixes and breakfast cereals": "GROCERY",

    # Then the top level, as the workbook names it. These strings are the
    # source's, not ours: check them against the file after an upstream
    # release, because a renamed category silently sends a whole shelf to
    # "OTHER" and the list still looks plausible.
    "Vegetables": "PRODUCE",
    "Fruit": "PRODUCE",
    "Meat and offal": "MEAT_FISH",
    "Sausages and cold meats": "MEAT_FISH",
    "Fish": "MEAT_FISH",
    "Milk and dairy products": "DAIRY",
    "Eggs": "DAIRY",
    "Plant based protein foods and alternaives to animal products": "DAIRY",
    "Bread, flakes and breakfast cereals": "BAKERY",
    "Cereal products, pulses and potatoes": "GROCERY",
    "Nuts, seeds and oleaginous fruit": "GROCERY",
    "Fats and oils": "GROCERY",
    "Sweets": "GROCERY",
    "Savoury snacks": "GROCERY",
    "Non-alcoholic beverages": "GROCERY",
    "Alcoholic beverages": "GROCERY",
    "Prepared dishes": "GROCERY",
    "Various": "GROCERY",
}


def normalise(value: str) -> str:
    """
    The form names are matched on: lowercase, unaccented, punctuation to spaces.

    The same rule is implemented in Java, in `FoodMatcher.normalise`, and the
    two are checked against each other by a test. If they disagree, a name is
    stored in one shape and looked up in another, and nothing matches.
    """
    # Ligatures first: NFD does not take "œ" apart.
    expanded = value.lower().replace("œ", "oe").replace("æ", "ae").replace("ß", "ss")
    folded = unicodedata.normalize("NFD", expanded)
    folded = "".join(c for c in folded if unicodedata.category(c) != "Mn")
    folded = re.sub(r"[^a-z0-9]+", " ", folded)
    return re.sub(r"\s+", " ", folded).strip()


def load(path, sheet_index=0):
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[sheet_index]]
    rows = sheet.iter_rows(values_only=True)
    # Two lines of title before the header.
    next(rows)
    next(rows)
    header = [str(h).strip() if h else "" for h in next(rows)]
    return header, list(rows)


def number(raw):
    """
    A value the source could not give is empty, and stays empty.

    The workbook also writes "<0.1" and "tr." for a trace. Both are a real
    measurement of "almost none", and reading them as zero is closer to the
    truth than dropping the food for having them.
    """
    if raw is None:
        return None
    text = str(raw).strip().replace(",", ".")
    if text in ("", "-", "n. b.", "n.b."):
        return None
    if text.startswith("<") or text in ("tr.", "tr"):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return None


def download(url, target):
    import requests

    target.parent.mkdir(parents=True, exist_ok=True)
    # The site refuses a bare client; it is a browser download page.
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=120)
    response.raise_for_status()
    target.write_bytes(response.content)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--from-cache", action="store_true")
    args = parser.parse_args()

    books = {}
    for locale, url in WORKBOOKS.items():
        path = CACHE / f"{locale}.xlsx"
        if not args.from_cache or not path.exists():
            print(f"downloading {locale}…", file=sys.stderr)
            download(url, path)
        books[locale] = load(path)

    header, rows = books["en"]
    index = {name: i for i, name in enumerate(header)}
    for column in list(MACROS) + list(MICRONUTRIENTS) + ["ID", "Name", "Synonyms", "Category"]:
        if column not in index:
            raise SystemExit(f"the workbook no longer has a column named {column!r}")

    OUT.mkdir(parents=True, exist_ok=True)
    foods_path = OUT / "foods.csv"
    names_path = OUT / "names.csv"

    micro_codes = [code for code, _ in MICRONUTRIENTS.values()]
    columns = ["ref", "category", "aisle", *MACROS.values(), *micro_codes]

    # The other languages, by id.
    localised = {}
    for locale in ("fr", "de"):
        book_header, book_rows = books[locale]
        id_at, name_at, synonyms_at = 0, 3, 4
        localised[locale] = {
            str(row[id_at]): (row[name_at], row[synonyms_at])
            for row in book_rows
            if row[id_at] is not None
        }

    written = 0
    with foods_path.open("w", newline="", encoding="utf-8") as foods_file, \
            names_path.open("w", newline="", encoding="utf-8") as names_file:
        foods = csv.writer(foods_file)
        names = csv.writer(names_file)
        foods.writerow(columns)
        names.writerow(["ref", "locale", "name", "normalised"])

        for row in rows:
            ref = str(row[index["ID"]]) if row[index["ID"]] is not None else None
            english = row[index["Name"]]
            if not ref or not english:
                continue

            kcal = number(row[index["Energy, kilocalories (kcal)"]])
            if kcal is None:
                # A food with no energy value cannot serve the one thing every
                # screen asks for. It is dropped rather than guessed at.
                continue

            category = str(row[index["Category"]] or "")
            top = category.split("/")[0].strip()
            aisle = AISLES.get(category.strip(), AISLES.get(top, "OTHER"))
            values = [number(row[index[column]]) for column in MACROS]
            micro = [number(row[index[column]]) for column in MICRONUTRIENTS]
            foods.writerow([
                ref, category, aisle,
                *["" if v is None else f"{v:g}" for v in values],
                *["" if v is None else f"{v:g}" for v in micro],
            ])
            written += 1

            for locale, name, synonyms in (
                ("en", english, row[index["Synonyms"]]),
                ("fr", *localised["fr"].get(ref, (None, None))[:1],
                 localised["fr"].get(ref, (None, None))[1]),
                ("de", *localised["de"].get(ref, (None, None))[:1],
                 localised["de"].get(ref, (None, None))[1]),
            ):
                for candidate in [name, *str(synonyms or "").split(";")]:
                    text = str(candidate or "").strip()
                    if not text or text == "None":
                        continue
                    key = normalise(text)
                    if key:
                        names.writerow([ref, locale, text, key])

    print(f"{written} foods → {foods_path}")
    print(f"names → {names_path}")


if __name__ == "__main__":
    main()
